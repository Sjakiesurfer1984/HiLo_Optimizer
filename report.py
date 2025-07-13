# report.py
from matplotlib.pylab import f
import matplotlib.pyplot as plt
from matplotlib.ticker import ScalarFormatter, StrMethodFormatter
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.formatting.rule import FormulaRule
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font
import yfinance as yf
from pathlib import Path
import pandas as pd


def plot_results(opt_df: pd.DataFrame, symbol: str, out_file: Path = None) -> Path:
    fig, ax = plt.subplots(figsize=(12, 7))
    # compute percentages
    gross_pct = (opt_df['Cumulative Return %'] - 1) * 100
    net_pct = (opt_df['Cumulative Return (net) %'] - 1) * 100
    fee_pct = gross_pct - net_pct
    x = opt_df['HiLo']
    # stacked bars: net return and fee drag
    ax.bar(x, net_pct, label='Net Return')
    ax.bar(x, fee_pct, bottom=net_pct,
           label='Gross Return (Trade fees excluded)',)
    ax.set_xlabel('HiLo Period')
    ax.set_ylabel('Cumulative Return (%)')
    ax.set_title(f'{symbol} HiLo Optimization')
    ax.legend()
    plt.tight_layout()

    ax.yaxis.set_major_formatter(StrMethodFormatter('{x:,.0f}'))
    # expand left margin so labels are not cut off
    fig.subplots_adjust(left=0.15)

    if out_file:
        fig.savefig(out_file)
        plt.close(fig)
        return out_file
    plt.show()
    plt.close(fig)
    return None


def plot_comparison(
    df: pd.DataFrame,
    best_period: int,
    symbol: str,
    initial_capital: float = 10000,
    out_file: Path = None
) -> Path:
    """
    Plot strategy vs buy & hold vs S&P 500, showing growth of initial capital.

    Args:
        df: Backtest DataFrame, must contain 'Cumulative Return' and 'Cumulative Return (net)'.
        best_period: The optimized HiLo period.
        symbol: Ticker symbol string.
        initial_capital: Starting capital in dollars.
        out_file: Optional Path to save the figure.

    Returns:
        The Path of the saved figure if out_file provided, else None.
    """
    fig, ax = plt.subplots(figsize=(12, 7))

    # compute growth of initial capital
    gross_val = df['Cumulative Return'] * initial_capital
    net_val = df['Cumulative Return (net)'] * initial_capital
    buy_val = (df['Adj Close'] / df['Adj Close'].iloc[0]) * initial_capital

    # plot growth curves
    ax.plot(df.index, gross_val,
            label=f'Strategy Gross (HiLo={best_period})')
    ax.plot(df.index, net_val,
            label=f'Strategy Net (HiLo={best_period})', linestyle='-')
    ax.plot(df.index, buy_val,
            label='Buy & Hold', linestyle='-')

    # fetch and plot S&P 500 growth
    sp = yf.download(
        tickers='^GSPC',
        start=df.index[0].strftime('%Y-%m-%d'),
        end=df.index[-1].strftime('%Y-%m-%d'),
        auto_adjust=False,
        group_by='column',
        progress=False
    )
    if isinstance(sp.columns, pd.MultiIndex):
        sp.columns = sp.columns.get_level_values(0)
    sp_adj = sp['Adj Close'].tz_localize(None).reindex(df.index).ffill()
    sp_val = (sp_adj / sp_adj.iloc[0]) * initial_capital
    ax.plot(df.index, sp_val, label='S&P 500', linestyle='-')

    # finalize plot
    ax.set_xlabel('Date')
    ax.set_ylabel(f'Value of ${initial_capital:,.0f} Investment')
    ax.set_title(
        f'{symbol} Strategy vs Benchmark (Growth of ${initial_capital:,.0f})')
    ax.legend()
    ax.grid(True)
    plt.tight_layout()

    ax.yaxis.set_major_formatter(StrMethodFormatter('{x:,.0f}'))
    # expand left margin so labels are not cut off
    fig.subplots_adjust(left=0.15)

    if out_file:
        fig.savefig(out_file)
        plt.close(fig)
        return out_file
    plt.show()
    plt.close(fig)
    return None


def plot_signals_with_returns(
    df: pd.DataFrame,
    best_period: int,
    symbol: str,
    out_file: Path = None
) -> Path:
    """
    Plot gross and net cumulative returns with buy arrows below the NET return line
    and sell arrows above the GROSS return line.
    """
    fig, ax = plt.subplots(figsize=(12, 7))

    # compute cumulative returns
    gross_pct = df['Cumulative Return'] * 100
    net_pct = df['Cumulative Return (net)'] * 100

    # plot gross and net lines
    ax.plot(df.index, gross_pct,
            label=f'Gross Return (HiLo={best_period})', linestyle='-')
    ax.plot(df.index, net_pct,
            label=f'Net Return   (HiLo={best_period})', linestyle='-')

    # identify position transitions to ensure alternating signals
    pos = df['Position']
    buy_idx = df.index[(pos == 1) & (pos.shift(1) != 1)]
    sell_idx = df.index[(pos == -1) & (pos.shift(1) != -1)]

    # calculate a small vertical offset (10% of the full gross range)
    y_range = gross_pct.max() - gross_pct.min()
    offset = y_range * 0.10

    # plot buy arrows below the net return line
    buy_y = net_pct.loc[buy_idx] - offset
    ax.scatter(buy_idx, buy_y,
               marker='^', color='green', s=100, label='Buy Signal')

    # plot sell arrows above the gross return line
    sell_y = gross_pct.loc[sell_idx] + offset
    ax.scatter(sell_idx, sell_y,
               marker='v', color='red', s=100, label='Sell Signal')

    # finalize
    ax.set_xlabel('Date')
    ax.set_ylabel('Cumulative Return (%)')
    ax.set_title(f'{symbol} Returns & Signals (HiLo={best_period})')
    ax.legend()
    ax.grid(True)
    plt.tight_layout()

    ax.yaxis.set_major_formatter(StrMethodFormatter('{x:,.0f}'))
    # expand left margin so labels are not cut off
    fig.subplots_adjust(left=0.15)

    if out_file:
        fig.savefig(out_file)
        plt.close(fig)
        return out_file
    plt.show()
    plt.close(fig)
    return None


def save_to_excel(df: pd.DataFrame, opt_df: pd.DataFrame, imgs: list, out_path: str):
    # 1) write the Data & Optimization sheets
    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Data', index=True)
        opt_df.to_excel(writer, sheet_name='Optimization', index=False)

    # 2) reopen with openpyxl
    wb = load_workbook(out_path)

    # 3) freeze the first row on every sheet
    for ws in wb.worksheets:
        ws.freeze_panes = 'A2'

    # 4) insert Plots sheet and pictures without overlap
    plots_ws = wb.create_sheet('Plots')
    col = 'A'
    row = 1
    for img_path in imgs:
        img = ExcelImage(str(img_path))
        plots_ws.add_image(img, f"{col}{row}")
        # calculate rows needed (approx. img.height / 14.5px per row)
        rows_to_skip = int(img.height / 14.5) - 7
        row += rows_to_skip

    # 5) add color scale to Optimization sheet
    opt_ws = wb['Optimization']
    last_row = opt_ws.max_row
    rule = ColorScaleRule(
        start_type='min', start_color='FFFF0000',
        mid_type='num', mid_value=1, mid_color='FFFFFFFF',
        end_type='max', end_color='FF00FF00'
    )
    # apply to 'Cumulative Return' in col C
    opt_ws.conditional_formatting.add(f'C2:C{last_row}', rule)

    # Number of net returns > 1x
    nr_pos_return = (opt_df['Cumulative Return (net) %'] > 1).sum()
    print(f"Number of net returns > 1x: {nr_pos_return}"
          )
    # Number of net returns < 1x
    nr_neg_return = (opt_df['Cumulative Return (net) %'] < 1).sum()
    print(f"Number of net returns < 1x: {nr_neg_return}")

    last_col = opt_ws.max_column
    col_pct_pos = get_column_letter(last_col + 1)
    col_pct_neg = get_column_letter(last_col + 2)
    opt_ws[f'{col_pct_pos}1'] = '% Positive'
    opt_ws[f'{col_pct_neg}1'] = '% Negative'
    opt_ws[f'{col_pct_pos}2'] = nr_pos_return
    opt_ws[f'{col_pct_neg}2'] = nr_neg_return

    # 6) add Total Cost and Trades to Data sheet
    data_ws = wb['Data']
    last_col = data_ws.max_column
    col_cost = get_column_letter(last_col + 2)
    col_trades = get_column_letter(last_col + 3)
    total_cost = df['Cost'].sum()

    # instead of diff().abs().sum(), do:
    buy_signals = ((df['Position'] == 1) & (
        df['Position'].shift(1) != 1)).sum()
    sell_signals = ((df['Position'] == -1) &
                    (df['Position'].shift(1) != -1)).sum()
    num_trades = int(buy_signals + sell_signals)

    data_ws[f'{col_cost}1'] = 'Total Cost'
    data_ws[f'{col_cost}2'] = total_cost
    data_ws[f'{col_trades}1'] = 'Trades'
    data_ws[f'{col_trades}2'] = num_trades
    bold = Font(bold=True)
    data_ws[f'{col_cost}1'].font = bold
    data_ws[f'{col_trades}1'].font = bold

    # 7) autofit all columns on every sheet
    for ws in wb.worksheets:
        for col_cells in ws.columns:
            col_letter = get_column_letter(col_cells[0].column)
            max_length = max((len(str(cell.value))
                             for cell in col_cells if cell.value is not None), default=0)
            ws.column_dimensions[col_letter].width = max_length + 2

    # 8) save
    print(f"Saving report to {out_path}")
    wb.save(out_path)
    print("Report saved successfully.")
